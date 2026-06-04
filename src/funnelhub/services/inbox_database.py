from __future__ import annotations

import csv
import io
import uuid
from dataclasses import dataclass
from datetime import datetime
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from sqlalchemy import Select, false, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from funnelhub.db.models import (
    Conversation,
    EmailSubscription,
    FunnelState,
    ImportBatch,
    ImportRow,
    Lead,
    LeadConsent,
    LeadContact,
    LeadCustomField,
    LeadExternalId,
    LeadUtm,
    Message,
    MessengerIdentity,
)
from funnelhub.services.getcourse_webhook import (
    ADDITIONAL_FIELD_LABELS,
    CONSENT_CUSTOM_FIELD_MAPPING,
    ingest_getcourse_webhook,
)

CSV_EXPORT_COLUMNS = [
    "lead_id",
    "getcourse_user_id",
    "name",
    "first_name",
    "last_name",
    "email",
    "phone",
    "city",
    "country",
    "source",
    "status",
    "telegram",
    "vk",
    "conversations_count",
    "messages_count",
    "created_at",
    "updated_at",
]
XLSX_EXPORT_COLUMNS = [
    ("lead_id", "ID FunnelHub"),
    ("getcourse_user_id", "ID GetCourse"),
    ("name", "Имя"),
    ("email", "Email"),
    ("phone", "Телефон"),
    ("country", "Страна"),
    ("city", "Город"),
    ("source", "Источник"),
    ("registration_type", "Тип регистрации"),
    ("getcourse_created_at", "Создан в GetCourse"),
    ("getcourse_last_activity_at", "Последняя активность GetCourse"),
    ("status", "Статус"),
    ("telegram", "Telegram"),
    ("vk", "VK"),
    ("utm_source", "utm_source"),
    ("utm_medium", "utm_medium"),
    ("utm_campaign", "utm_campaign"),
    ("utm_term", "utm_term"),
    ("utm_content", "utm_content"),
    ("utm_group", "utm_group"),
    ("vk_id", "VK-ID из GetCourse"),
    ("getcourse_groups", "Группы GetCourse"),
    ("partner", "От партнера"),
    ("partner_id", "ID партнера"),
    ("partner_email", "Email партнера"),
    ("partner_name", "ФИО партнера"),
    ("manager_name", "ФИО менеджера"),
    ("consents", "Согласия"),
    ("custom_fields", "Дополнительные поля"),
    ("conversations_count", "Диалогов"),
    ("messages_count", "Сообщений"),
    ("created_at", "Создан в FunnelHub"),
    ("updated_at", "Обновлен в FunnelHub"),
]

IMPORT_FIELD_ALIASES = {
    "gc_user_id": ("gc_user_id", "getcourse_user_id", "id", "ID", "GetCourse ID"),
    "name": ("name", "full_name", "ФИО", "Name"),
    "first_name": ("first_name", "Имя", "Имя пользователя", "First name"),
    "last_name": ("last_name", "Фамилия", "Last name"),
    "email": ("email", "Email", "E-mail", "Почта", "Эл. почта"),
    "phone": ("phone", "Телефон", "Phone", "Номер телефона"),
    "city": ("city", "Город", "City"),
    "country": ("country", "Страна", "Country"),
    "source": ("source", "Источник", "Source", "Откуда пришел"),
    "registration_type": ("registration_type", "Тип регистрации"),
    "created": ("created", "created_at", "Создан"),
    "last_activity": ("last_activity", "last_activity_at", "Последняя активность"),
    "utm_source": ("utm_source",),
    "utm_medium": ("utm_medium",),
    "utm_campaign": ("utm_campaign",),
    "utm_term": ("utm_term",),
    "utm_content": ("utm_content",),
    "utm_group": ("utm_group",),
    "vk_id": ("vk_id", "VK-ID"),
    "getcourse_groups": ("getcourse_groups", "id групп пользователя/дата добавления"),
}
GETCOURSE_EXPORT_CONSENT_FIELD_KEYS = tuple(CONSENT_CUSTOM_FIELD_MAPPING)


@dataclass(frozen=True)
class DatabaseLeadSummary:
    id: uuid.UUID
    getcourse_user_id: int | None
    name: str | None
    email: str | None
    phone: str | None
    city: str | None
    country: str | None
    source: str | None
    status: str
    created_at: datetime
    updated_at: datetime
    telegram: str | None
    vk: str | None
    conversations_count: int
    messages_count: int


@dataclass(frozen=True)
class DatabaseLeadDetail:
    lead: DatabaseLeadSummary
    profile_fields: list[dict[str, Any]]
    contacts: list[dict[str, Any]]
    identities: list[dict[str, Any]]
    external_ids: list[dict[str, Any]]
    utm_snapshots: list[dict[str, Any]]
    custom_fields: list[dict[str, Any]]
    consents: list[dict[str, Any]]
    email_subscriptions: list[dict[str, Any]]
    funnel_states: list[dict[str, Any]]
    recent_messages: list[dict[str, Any]]
    raw_getcourse_data: dict[str, Any]


@dataclass(frozen=True)
class DatabaseLeadList:
    items: list[DatabaseLeadSummary]
    total: int
    limit: int
    offset: int


@dataclass(frozen=True)
class DatabaseImportResult:
    batch_id: uuid.UUID
    total_rows: int
    processed_rows: int
    failed_rows: int
    created_rows: int
    updated_rows: int
    errors: list[dict[str, Any]]


async def list_database_leads(
    session: AsyncSession,
    *,
    query: str | None = None,
    limit: int = 50,
    offset: int = 0,
) -> DatabaseLeadList:
    clean_query = query.strip() if query else None
    base_statement = build_lead_summary_query()
    count_statement = select(func.count()).select_from(Lead)

    if clean_query:
        search_filter = build_lead_search_filter(clean_query)
        base_statement = base_statement.where(search_filter)
        count_statement = count_statement.where(search_filter)

    total = int(await session.scalar(count_statement) or 0)
    rows = (
        await session.execute(
            base_statement.order_by(Lead.created_at.desc()).limit(limit).offset(offset)
        )
    ).all()
    return DatabaseLeadList(
        items=[row_to_lead_summary(row) for row in rows],
        total=total,
        limit=limit,
        offset=offset,
    )


async def get_database_lead_detail(
    session: AsyncSession,
    lead_id: uuid.UUID,
) -> DatabaseLeadDetail | None:
    row = (
        await session.execute(build_lead_summary_query().where(Lead.id == lead_id))
    ).one_or_none()
    if row is None:
        return None

    contacts = (
        await session.scalars(
            select(LeadContact)
            .where(LeadContact.lead_id == lead_id)
            .order_by(LeadContact.created_at)
        )
    ).all()
    identities = (
        await session.scalars(
            select(MessengerIdentity)
            .where(MessengerIdentity.lead_id == lead_id)
            .order_by(MessengerIdentity.channel)
        )
    ).all()
    external_ids = (
        await session.scalars(
            select(LeadExternalId)
            .where(LeadExternalId.lead_id == lead_id)
            .order_by(LeadExternalId.provider)
        )
    ).all()
    utm_snapshots = (
        await session.scalars(
            select(LeadUtm)
            .where(LeadUtm.lead_id == lead_id, LeadUtm.source_kind != "getcourse_system")
            .order_by(LeadUtm.created_at.desc())
        )
    ).all()
    custom_fields = (
        await session.scalars(
            select(LeadCustomField)
            .where(LeadCustomField.lead_id == lead_id)
            .order_by(LeadCustomField.field_position, LeadCustomField.field_key)
        )
    ).all()
    consents = (
        await session.scalars(
            select(LeadConsent)
            .where(LeadConsent.lead_id == lead_id)
            .order_by(LeadConsent.consent_type)
        )
    ).all()
    email_subscriptions = (
        await session.scalars(
            select(EmailSubscription)
            .where(EmailSubscription.lead_id == lead_id)
            .order_by(EmailSubscription.created_at.desc())
        )
    ).all()
    funnel_states = (
        await session.scalars(
            select(FunnelState)
            .where(FunnelState.lead_id == lead_id)
            .order_by(FunnelState.created_at)
        )
    ).all()
    messages = (
        await session.scalars(
            select(Message)
            .where(Message.lead_id == lead_id)
            .order_by(Message.created_at.desc())
            .limit(20)
        )
    ).all()

    lead = await session.get(Lead, lead_id)
    if lead is None:
        return None

    return DatabaseLeadDetail(
        lead=row_to_lead_summary(row),
        profile_fields=lead_profile_fields(lead),
        contacts=[
            {
                "type": contact.contact_type,
                "value": contact.value,
                "is_primary": contact.is_primary,
                "is_verified": contact.is_verified,
            }
            for contact in contacts
        ],
        identities=[
            {
                "channel": identity.channel,
                "external_user_id": identity.external_user_id,
                "username": identity.username,
                "display_name": identity.display_name,
                "is_subscribed": identity.is_subscribed,
            }
            for identity in identities
        ],
        external_ids=[
            {
                "provider": item.provider,
                "external_id": item.external_id,
            }
            for item in external_ids
        ],
        utm_snapshots=[
            {
                "source_kind": item.source_kind,
                "utm_source": item.utm_source,
                "utm_medium": item.utm_medium,
                "utm_campaign": item.utm_campaign,
                "utm_term": item.utm_term,
                "utm_content": item.utm_content,
                "utm_group": item.utm_group,
                "created_at": item.created_at,
            }
            for item in utm_snapshots
        ],
        custom_fields=[
            {
                "key": field.field_key,
                "label": field.field_label or human_field_label(field.field_key),
                "value": field.value,
                "normalized_bool": field.normalized_bool,
                "position": field.field_position,
            }
            for field in custom_fields
        ],
        consents=[
            {
                "type": consent.consent_type,
                "is_granted": consent.is_granted,
                "granted_at": consent.granted_at,
                "revoked_at": consent.revoked_at,
                "metadata": consent.metadata_ or {},
            }
            for consent in consents
        ],
        email_subscriptions=[
            {
                "email": subscription.email,
                "status": subscription.status,
                "provider": subscription.provider,
                "subscribed_at": subscription.subscribed_at,
                "unsubscribed_at": subscription.unsubscribed_at,
            }
            for subscription in email_subscriptions
        ],
        funnel_states=[
            {
                "funnel_key": state.funnel_key,
                "status": state.status,
                "current_step_key": state.current_step_key,
                "next_run_at": state.next_run_at,
                "completed_at": state.completed_at,
            }
            for state in funnel_states
        ],
        recent_messages=[
            {
                "id": str(message.id),
                "channel": message.channel,
                "direction": message.direction,
                "body": message.body,
                "status": message.status,
                "created_at": message.created_at,
            }
            for message in messages
        ],
        raw_getcourse_data=lead.raw_getcourse_data or {},
    )


async def export_database_leads_csv(session: AsyncSession, *, query: str | None = None) -> str:
    lead_list = await list_database_leads(session, query=query, limit=10_000, offset=0)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_EXPORT_COLUMNS, lineterminator="\n")
    writer.writeheader()
    for lead in lead_list.items:
        writer.writerow(lead_export_row(lead))
    return output.getvalue()


async def export_database_leads_xlsx(session: AsyncSession, *, query: str | None = None) -> bytes:
    lead_list = await list_database_leads(session, query=query, limit=10_000, offset=0)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Лиды"
    headers = [label for _, label in XLSX_EXPORT_COLUMNS]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="24483C")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for lead in lead_list.items:
        detail = await get_database_lead_detail(session, lead.id)
        if detail is None:
            continue
        export_row = xlsx_export_row(detail)
        sheet.append([xlsx_cell_value(export_row.get(key, "")) for key, _ in XLSX_EXPORT_COLUMNS])

    for column_cells in sheet.columns:
        first_cell = column_cells[0]
        column_letter = first_cell.column_letter
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
        )
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


async def import_database_leads_csv(
    session: AsyncSession,
    *,
    file_name: str,
    content: bytes,
) -> DatabaseImportResult:
    text = decode_csv_content(content)
    delimiter = detect_csv_delimiter(text)
    csv_reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        fieldnames = next(csv_reader)
    except StopIteration:
        raise ValueError("CSV file must contain a header row.") from None
    if not fieldnames:
        raise ValueError("CSV file must contain a header row.")

    batch = ImportBatch(
        id=uuid.uuid4(),
        source="inbox",
        file_name=file_name,
        file_format="csv",
        encoding="utf-8-sig",
        delimiter=delimiter,
        status="processing",
        metadata_={"fieldnames": fieldnames},
    )
    session.add(batch)
    await session.flush()

    total_rows = 0
    processed_rows = 0
    failed_rows = 0
    created_rows = 0
    updated_rows = 0
    errors: list[dict[str, Any]] = []

    for row_number, values in enumerate(csv_reader, start=2):
        total_rows += 1
        row = import_values_to_row(fieldnames, values)
        payload = import_row_to_payload(row)
        import_row = ImportRow(
            id=uuid.uuid4(),
            batch_id=batch.id,
            row_number=row_number,
            raw_data={key: value for key, value in row.items() if key is not None},
            errors=[],
        )
        session.add(import_row)

        try:
            if not has_import_identity(payload):
                raise ValueError("Row must include gc_user_id, email, or phone.")
            result = await ingest_getcourse_webhook(session, payload)
            import_row.lead_id = result.lead_id
            import_row.status = "imported"
            processed_rows += 1
            if result.created:
                created_rows += 1
            else:
                updated_rows += 1
        except ValueError as exc:
            failed_rows += 1
            error = {"row_number": row_number, "message": str(exc)}
            import_row.status = "failed"
            import_row.errors = [error]
            errors.append(error)

    batch.total_rows = total_rows
    batch.processed_rows = processed_rows
    batch.failed_rows = failed_rows
    batch.status = "completed" if processed_rows > 0 else "failed"
    batch.metadata_ = {
        **(batch.metadata_ or {}),
        "created_rows": created_rows,
        "updated_rows": updated_rows,
    }
    return DatabaseImportResult(
        batch_id=batch.id,
        total_rows=total_rows,
        processed_rows=processed_rows,
        failed_rows=failed_rows,
        created_rows=created_rows,
        updated_rows=updated_rows,
        errors=errors[:20],
    )


def build_lead_summary_query() -> Select[tuple[Any, ...]]:
    email = lead_contact_subquery("email")
    phone = lead_contact_subquery("phone")
    telegram = messenger_identity_subquery("telegram")
    vk = messenger_identity_subquery("vk")
    conversations_count = (
        select(func.count(Conversation.id))
        .where(Conversation.lead_id == Lead.id)
        .correlate(Lead)
        .scalar_subquery()
    )
    messages_count = (
        select(func.count(Message.id))
        .where(Message.lead_id == Lead.id)
        .correlate(Lead)
        .scalar_subquery()
    )
    return select(
        Lead.id,
        Lead.getcourse_user_id,
        Lead.full_name,
        Lead.first_name,
        Lead.last_name,
        Lead.city,
        Lead.country,
        Lead.source,
        Lead.status,
        Lead.created_at,
        Lead.updated_at,
        email.label("email"),
        phone.label("phone"),
        telegram.label("telegram"),
        vk.label("vk"),
        conversations_count.label("conversations_count"),
        messages_count.label("messages_count"),
    )


def build_lead_search_filter(query: str) -> Any:
    like_query = f"%{query}%"
    lead_filters: list[Any] = [
        Lead.full_name.ilike(like_query),
        Lead.first_name.ilike(like_query),
        Lead.last_name.ilike(like_query),
        Lead.source.ilike(like_query),
    ]
    if query.isdigit():
        lead_filters.append(Lead.getcourse_user_id == int(query))
    contact_match = select(LeadContact.lead_id).where(LeadContact.value.ilike(like_query))
    identity_match = select(MessengerIdentity.lead_id).where(
        or_(
            MessengerIdentity.username.ilike(like_query),
            MessengerIdentity.display_name.ilike(like_query),
            MessengerIdentity.external_user_id.ilike(like_query),
        )
    )
    return or_(
        Lead.id.in_(contact_match),
        Lead.id.in_(identity_match),
        *lead_filters,
        false(),
    )


def lead_contact_subquery(contact_type: str) -> Any:
    return (
        select(LeadContact.value)
        .where(LeadContact.lead_id == Lead.id, LeadContact.contact_type == contact_type)
        .order_by(LeadContact.is_primary.desc(), LeadContact.created_at.asc())
        .limit(1)
        .correlate(Lead)
        .scalar_subquery()
    )


def messenger_identity_subquery(channel: str) -> Any:
    return (
        select(
            func.coalesce(
                MessengerIdentity.username,
                MessengerIdentity.display_name,
                MessengerIdentity.external_user_id,
            )
        )
        .where(
            MessengerIdentity.lead_id == Lead.id,
            MessengerIdentity.channel == channel,
            MessengerIdentity.is_subscribed.is_(True),
        )
        .order_by(MessengerIdentity.created_at.desc())
        .limit(1)
        .correlate(Lead)
        .scalar_subquery()
    )


def row_to_lead_summary(row: Any) -> DatabaseLeadSummary:
    name = row.full_name or " ".join(part for part in [row.first_name, row.last_name] if part)
    return DatabaseLeadSummary(
        id=row.id,
        getcourse_user_id=row.getcourse_user_id,
        name=name or None,
        email=row.email,
        phone=row.phone,
        city=row.city,
        country=row.country,
        source=row.source,
        status=row.status,
        created_at=row.created_at,
        updated_at=row.updated_at,
        telegram=row.telegram,
        vk=row.vk,
        conversations_count=int(row.conversations_count or 0),
        messages_count=int(row.messages_count or 0),
    )


def lead_export_row(lead: DatabaseLeadSummary) -> dict[str, Any]:
    return {
        "lead_id": str(lead.id),
        "getcourse_user_id": lead.getcourse_user_id,
        "name": lead.name,
        "first_name": "",
        "last_name": "",
        "email": lead.email,
        "phone": lead.phone,
        "city": lead.city,
        "country": lead.country,
        "source": lead.source,
        "status": lead.status,
        "telegram": lead.telegram,
        "vk": lead.vk,
        "conversations_count": lead.conversations_count,
        "messages_count": lead.messages_count,
        "created_at": lead.created_at.isoformat(),
        "updated_at": lead.updated_at.isoformat(),
    }


def xlsx_export_row(detail: DatabaseLeadDetail) -> dict[str, Any]:
    lead = detail.lead
    profile = detail_profile_dict(detail.profile_fields)
    custom = detail_custom_field_dict(detail.custom_fields)
    form_utm = latest_utm_snapshot(detail.utm_snapshots, "form")
    return {
        "lead_id": str(lead.id),
        "getcourse_user_id": lead.getcourse_user_id,
        "name": lead.name,
        "email": lead.email,
        "phone": lead.phone,
        "country": lead.country,
        "city": lead.city,
        "source": lead.source,
        "registration_type": profile.get("registration_type"),
        "getcourse_created_at": profile.get("getcourse_created_at"),
        "getcourse_last_activity_at": profile.get("getcourse_last_activity_at"),
        "status": lead.status,
        "telegram": lead.telegram,
        "vk": lead.vk,
        "utm_source": form_utm.get("utm_source"),
        "utm_medium": form_utm.get("utm_medium"),
        "utm_campaign": form_utm.get("utm_campaign"),
        "utm_term": form_utm.get("utm_term"),
        "utm_content": form_utm.get("utm_content"),
        "utm_group": form_utm.get("utm_group"),
        "vk_id": custom.get("vk_id"),
        "getcourse_groups": custom.get("getcourse_groups"),
        "partner": custom.get("partner"),
        "partner_id": custom.get("partner_id"),
        "partner_email": custom.get("partner_email"),
        "partner_name": custom.get("partner_name"),
        "manager_name": custom.get("manager_name"),
        "consents": "\n".join(format_consent(consent) for consent in detail.consents),
        "custom_fields": "\n".join(format_custom_field(field) for field in detail.custom_fields),
        "conversations_count": lead.conversations_count,
        "messages_count": lead.messages_count,
        "created_at": lead.created_at.isoformat(),
        "updated_at": lead.updated_at.isoformat(),
    }


def xlsx_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def decode_csv_content(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV file encoding must be UTF-8 or Windows-1251.")


def detect_csv_delimiter(text: str) -> str:
    header = text.splitlines()[0] if text.splitlines() else ""
    candidates = [",", "\t", ";"]
    return max(candidates, key=lambda delimiter: header.count(delimiter))


def import_values_to_row(fieldnames: list[str], values: list[str]) -> dict[str, str | None]:
    row: dict[str, str | None] = {}
    headerless_custom_index = 0
    for index, value in enumerate(values):
        header = fieldnames[index] if index < len(fieldnames) else ""
        key = import_column_key(header, index, headerless_custom_index)
        if not header.strip():
            headerless_custom_index += 1
        row[key] = value
    return row


def import_column_key(header: str, index: int, headerless_custom_index: int) -> str:
    cleaned = header.strip()
    if cleaned:
        return cleaned
    if headerless_custom_index < len(GETCOURSE_EXPORT_CONSENT_FIELD_KEYS):
        return GETCOURSE_EXPORT_CONSENT_FIELD_KEYS[headerless_custom_index]
    return f"column_{index + 1}"


def import_row_to_payload(row: dict[str, str | None]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for target_key, aliases in IMPORT_FIELD_ALIASES.items():
        value = first_row_value(row, aliases)
        if value is not None:
            payload[target_key] = value

    for key, value in row.items():
        if key is not None and key.startswith("custom_") and clean_import_value(value) is not None:
            payload[key] = clean_import_value(value)

    return payload


def first_row_value(row: dict[str, str | None], keys: tuple[str, ...]) -> str | None:
    for key in keys:
        value = clean_import_value(row.get(key))
        if value is not None:
            return value
    return None


def clean_import_value(value: str | None) -> str | None:
    if value is None:
        return None
    cleaned = value.strip()
    return cleaned or None


def has_import_identity(payload: dict[str, Any]) -> bool:
    return any(payload.get(key) for key in ("gc_user_id", "email", "phone"))


def lead_profile_fields(lead: Lead) -> list[dict[str, Any]]:
    fields = [
        ("getcourse_user_id", "ID GetCourse", lead.getcourse_user_id),
        ("full_name", "Имя", lead.full_name),
        ("first_name", "Имя пользователя", lead.first_name),
        ("last_name", "Фамилия", lead.last_name),
        ("country", "Страна", lead.country),
        ("city", "Город", lead.city),
        ("source", "Источник", lead.source),
        ("registration_type", "Тип регистрации", lead.registration_type),
        ("getcourse_created_at", "Создан в GetCourse", lead.getcourse_created_at),
        (
            "getcourse_last_activity_at",
            "Последняя активность GetCourse",
            lead.getcourse_last_activity_at,
        ),
        ("created_at", "Создан в FunnelHub", lead.created_at),
        ("updated_at", "Обновлен в FunnelHub", lead.updated_at),
    ]
    return [
        {"key": key, "label": label, "value": value}
        for key, label, value in fields
        if value is not None
    ]


def detail_profile_dict(fields: list[dict[str, Any]]) -> dict[str, Any]:
    return {str(field["key"]): field["value"] for field in fields}


def detail_custom_field_dict(fields: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        str(field["key"]): field["value"]
        for field in fields
        if field.get("value") is not None
    }


def latest_utm_snapshot(snapshots: list[dict[str, Any]], source_kind: str) -> dict[str, Any]:
    for snapshot in snapshots:
        if snapshot.get("source_kind") == source_kind:
            return snapshot
    return {}


def format_consent(consent: dict[str, Any]) -> str:
    status = "да" if consent.get("is_granted") else "нет"
    return f"{human_consent_label(str(consent.get('type')))}: {status}"


def format_custom_field(field: dict[str, Any]) -> str:
    label = str(field.get("label") or field.get("key"))
    value = field.get("value")
    if field.get("normalized_bool") is True:
        value = "да"
    elif field.get("normalized_bool") is False:
        value = "нет"
    return f"{label}: {value}"


def human_consent_label(consent_type: str) -> str:
    labels = {
        "privacy_policy": "Политика конфиденциальности",
        "offer_agreement": "Договор оферты",
        "personal_data": "Обработка персональных данных",
        "email_marketing": "Email-рассылки",
        "messenger_marketing": "Рассылки в мессенджерах",
    }
    return labels.get(consent_type, consent_type)


def human_field_label(field_key: str) -> str:
    return ADDITIONAL_FIELD_LABELS.get(field_key, field_key)
